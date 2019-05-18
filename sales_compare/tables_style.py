from mouth_oder import split_table

import glob, os, openpyxl
from openpyxl.styles import PatternFill,Border, Side, Alignment, Font
from openpyxl.styles.colors import RED,YELLOW, BLUE, BLACK, GREEN


# 调用start方法，将文件拆分到指定的文件夹中
split_table.start()

print('正在设置表格的样式，请耐心等待。。。')
# 添加表格的格式
split_table = split_table.Split_Table()
base_path = split_table.son_name()

# 字体以及填充参数
alignment_center=Alignment(horizontal='center', vertical='center')
alignment_left=Alignment(horizontal='left', vertical='center')

thin = Side(border_style="thin",color=BLACK)
border = Border(top=thin, left=thin, right=thin, bottom=thin)
row_title_font = Font(name='黑体', size=12, bold=True, color=BLACK)
column_title_font = Font(name='黑体', size=12, color=BLACK)
row_title_fill = PatternFill(fill_type='solid',fgColor='FFCC33')
column_title_fill = PatternFill(fill_type='solid',fgColor='FFCC33')
content_font = Font(name='宋体', size=11, bold=False,color=BLACK)
content_fill = PatternFill(fill_type='solid' ,fgColor=YELLOW)
title_font = Font(name="宋体", bold=True, size=24)


# 找到所有的表格文件路径并打开
files_path = glob.glob(base_path + '*.xlsx')
for file_path in files_path:
    filepath, fullflname = os.path.split(file_path)
    fname, ext = os.path.splitext(fullflname)
    wb = openpyxl.load_workbook(file_path)
    ws = wb.get_sheet_by_name(fname)
    print(ws)

    # 锁定首行
    ws.freeze_panes = ws['A2']

    # 单元格样式
    for columns in ws.columns:
        i = len(columns)
        for n in range(len(columns) + 1):
            ws.row_dimensions[n].height = 15

        for cell in columns:
            cell.alignment = alignment_center
            cell.border = border
            if cell.column == 2 or cell.column == 3 or cell.column == 4 or cell.column == 5 or cell.column == 13:
                cell.alignment = alignment_left

    # 局部表头居中
    ws['B1'].alignment = alignment_center
    ws['C1'].alignment = alignment_center
    ws['D1'].alignment = alignment_center
    ws['E1'].alignment = alignment_center
    ws['M1'].alignment = alignment_center

    # 自动换行
    ws['H1'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
    ws['I1'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
    ws['J1'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
    ws['K1'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
    ws['M1'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')

    # 表头的填充
    ws['H1'].fill = PatternFill(fill_type='solid', fgColor=RED)
    ws['I1'].fill = PatternFill(fill_type='solid', fgColor=RED)
    ws['K1'].fill = PatternFill(fill_type='solid', fgColor=YELLOW)
    ws['L1'].fill = PatternFill(fill_type='solid', fgColor=YELLOW)

    ws.row_dimensions[1].height = 40
    ws.column_dimensions['A'].width = 7.5
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 13
    ws.column_dimensions['E'].width = 37
    ws.column_dimensions['F'].width = 5
    ws.column_dimensions['G'].width = 8.5
    ws.column_dimensions['H'].width = 10
    ws.column_dimensions['I'].width = 10
    ws.column_dimensions['J'].width = 11
    ws.column_dimensions['K'].width = 11
    ws.column_dimensions['L'].width = 9.75
    ws.column_dimensions['M'].width = 24

    # 备注的大小设置
    ws['M1'].font = Font(name='黑体', size=9, color=RED)

    # 保存结果
    wb.save(base_path + "%s.xlsx" % fname)